#	Script written on Python 3.6
# 	Author : Parag Fulzele
#	Description : Convert shelter old database data into XML files
#

import psycopg2
import xlrd
import openpyxl
import uuid
import datetime
import dicttoxml
import xml.etree.ElementTree as ET
import os
from time import gmtime, strftime
import traceback
import copy

#local
#db_host = "192.168.0.64"
#db_port = "5432"
#db_name = "shelterlive"
#db_user = "postgres"
#db_pwd = "softcorner"

##Live
db_host = "176.58.119.87"
db_port = ""
db_name = "sheltersurvey"
db_user = "shelter"
db_pwd = "XIcfe4cIV7ooGiIb"

question_map_dict = {}
question_option_map_dict = {}

option_dict = {}
city_ward_slum_dict = {}

gl_ra_xml_dict = {
	'formhub' : {
		'uuid' : None
	},
	'start' : None,
	'end' : None,
	'group_ws5ux48' : {
		'date_of_survey' : None,
		'name_of_surveyor_s_who_checke' : None,
		'name_of_surveyor_s_who_collec' : None,
		'name_of_surveyor_s_who_took_t' : None
	},
	'group_zl6oo94' : {
		'group_uj8eg07' : {
			'city' : None,
			'admin_ward' : None,
			'slum_name' : None,
			'survey_sector_number' : None,
			'landmark' : None
		},
		'group_wb1hp47' : {
			'year_established_according_to' : None,
			'legal_status' : None,
			'Date_of_declaration' : None,
			'land_owner' : None,
			'development_plan_reservation_t' : None,
			'development_plan_reservation' : None,
			'approximate_area_of_the_settle' : None,
			'number_of_huts_in_settlement' : None,
			'location' : None,
			'topography' : None,
			'describe_the_slum' : None
		}
	},
	'group_te3dx03' : {
		'group_ul75r92' : {
			'status_of_defecation' : None,
			'number_of_community_toilet_blo' : None,
			'number_of_pay_and_use_CTBs' : None
		},
		'group_tb3th42' : {
			'is_the_CTB_in_use' : None,
			'fee_for_use_of_ctb_per_family' : None,
			'cost_of_pay_and_use_toilet_pe' : None,
			'ctb_gender_usage' : None,
			'total_number_of_mixed_seats_al' : None,
			'number_of_mixed_seats_allotted' : None,
			'the_reason_for_the_mixed_seats' : None,
			'number_of_seats_allotted_to_me' : None,
			'number_of_seats_allotted_to_me_001' : None,
			'the_reason_for_men_not_using_t' : None,
			'number_of_seats_allotted_to_wo' : None,
			'number_of_seats_allotted_to_wo_001' : None,
			'the_reason_for_women_not_using' : None,
			'is_the_ctb_available_at_night' : None,
			'ctb_maintenance_provided_by' : None,
			'condition_of_ctb_structure' : None,
			'out_of_total_seats_no_of_pans_in_good_condition' : None,
			'out_of_total_seats_no_of_doors_in_good_condition' : None,
			'out_of_total_seats_no_of_seats_where_electricity_is_available' : None,
			'out_of_total_seats_no_of_seats_where_tiles_on_wall_are_in_good_condition' : None,
			'out_of_total_seats_no_of_seats_where_tiles_on_floor_are_in_good_condition' : None,
			'frequency_of_ctb_cleaning_by_U' : None,
			'does_the_ulb_ngo_communty_use' : None,
			'cleanliness_of_the_ctb' : None,
			'is_there_a_caretaker_for_the_C' : None,
			'type_of_water_supply_in_ctb' : None,
			'capacity_of_ctb_water_tank_in' : None,
			'litres_of_water_used_by_commun' : None,
			'availability_of_water_in_the_t' : None,
			'availability_of_electricity_in' : None,
			'availability_of_electricity_in_001' : None,
			'facility_in_the_toilet_block_f' : None,
			'condition_of_facility_for_chil' : None,
			'sewage_disposal_system' : None,
			'distance_to_nearest_ulb_sewer' : None
		},
		'toilet_comment' : None
	},
	'group_zj8tc43' : {
		'Total_number_of_standposts_in_' : None,
		'Total_number_of_standposts_NOT' : None,
		'total_number_of_taps_in_use_n' : None,
		'total_number_of_taps_in_use_n_001' : None,
		'total_number_of_handpumps_in_u' : None,
		'total_number_of_handpumps_in_u_001' : None,
		'alternative_source_of_water' : None,
		'availability_of_water' : None,
		'pressure_of_water_in_the_syste' : None,
		'coverage_of_wateracross_settle' : None,
		'quality_of_water_in_the_system' : None,
		'water_supply_comment' : None
	},
	'group_ks0wh10' : {
		'total_number_of_waste_containe' : None,
		'facility_of_waste_collection' : None,
		'frequency_of_waste_collection' : None,
		'frequency_of_waste_collection__002' : None,
		'frequency_of_waste_collection_001' : None,
		'frequency_of_waste_collection_' : None,
		'frequency_of_waste_collection__001' : None,
		'coverage_of_waste_collection_a' : None,
		'coverage_of_waste_collection_a_001' : None,
		'coverage_of_waste_collection_a_002' : None,
		'coverage_of_waste_collection_a_003' : None,
		'where_are_the_communty_open_du' : None,
		'do_the_member_of_community_dep' : None,
		'Waste_management_comments' : None
	},
	'group_kk5gz02' : {
		'presence_of_drains_within_the' : None,
		'coverage_of_drains_across_the' : None,
		'do_the_drains_get_blocked' : None,
		'is_the_drainage_gradient_adequ' : None,
		'diameter_of_ulb_sewer_line_acr' : None,
		'drainage_comment' : None
	},
	'group_bv7hf31' : {
		'Presence_of_gutter' : None,
		'type_of_gutter_within_the_sett' : None,
		'coverage_of_gutter' : None,
		'are_gutter_covered' : None,
		'do_gutters_flood' : None,
		'do_gutter_get_choked' : None,
		'is_gutter_gradient_adequate' : None,
		'comments_on_gutter' : None
	},
	'group_xy9hz30' : {
		'presence_of_roads_within_the_s' : None,
		'type_of_roads_within_the_settl' : None,
		'coverage_of_pucca_road_across' : None,
		'finish_of_the_road' : None,
		'average_width_of_internal_road' : None,
		'average_width_of_arterial_road' : None,
		'point_of_vehicular_access_to_t' : None,
		'is_the_settlement_below_or_abo' : None,
		'are_the_huts_below_or_above_th' : None,
		'road_and_access_comment' : None
	},
	'__version__' : None,
	'meta' : {
		'instanceID' : None
	}
}

gl_rhs_xml_dict = {
	'formhub' : {
		'uuid' : None
	},
	'start' : None,
	'end' : None,
	'group_ce0hf58' : {
		'city' : None,
		'admin_ward' : None,
		'slum_name' : None,
		'date_of_rhs' : None,
		'name_of_surveyor_who_collected_rhs_data' : None,
		'house_no' : None,
		'Type_of_structure_occupancy' : None
	},
	'group_ye18c77' : {
		'group_ud4em45' : {
			'what_is_the_full_name_of_the_family_head_' : None,
			'mobile_number' : None,
			'adhar_card_number' : None
		},
		'group_yw8pj39' : {
			'what_is_the_structure_of_the_house' : None,
			'what_is_the_ownership_status_of_the_house' : None,
			'number_of_family_members' : None,
			'Do_you_have_a_girl_child_under' : None,
			'if_yes_how_many_' : None,
			'house_area_in_sq_ft' : None,
			'Current_place_of_defecation_toilet' : None,
			'does_any_member_of_your_family_go_for_open_defecation_' : None,
			'where_the_individual_toilet_is_connected_to_' : None,
			'type_of_water_connection' : None,
			'facility_of_waste_collection' : None,
			'Are_you_interested_in_individu' : None,
			'if_yes_why_' : None,
			'if_no_why_' : None,
			'type_of_toilet_preference' : None,
			'Have_you_applied_for_indiviual' : None,
			'How_many_installements_have_yo' : None,
			'when_did_you_receive_the_first_installment_date' : None,
			'when_did_you_receive_the_second_installment_date' : None,
			'what_is_the_status_of_toilet_under_sbm_' : None,
			'Does_any_family_members_has_co' : None
		},
	},
	'__version__' : None,
	'meta' : {
		'instanceID' : None
	}
}

gl_ff_xml_dict = {
	'formhub' : {
		'uuid' : None
	},
	'group_vq77l17' : {
		'city' : None,
		'admin_ward' : None,
		'slum_name' : None,
		'Settlement_address' : None,
		'Household_number' : None
	},
	'group_oh4zf84' : {
		'Name_of_the_family_head' : None,
		'Name_of_Native_villa_district_and_state' : None,
		'Duration_of_stay_in_the_city' : None,
		'Duration_of_stay_in_s_current_settlement' : None,
		'Type_of_house' : None,
		'Ownership_status' : None
	},
	'group_im2th52' : {
		'Total_family_members' : None,
		'Number_of_Male_members' : None,
		'Number_of_Female_members' : None,
		'Number_of_Children_under_5_years_of_age' : None,
		'Number_of_members_over_60_years_of_age' : None,
		'Number_of_Girl_children_between_0_18_yrs' : None,
		'Number_of_disabled_members' : None,
		'If_yes_specify_type_of_disability' : None,
		'Number_of_earning_members' : None,
		'Occupation_s_of_earning_members' : None,
		'Approximate_monthly_family_income' : None
	},
	'group_ne3ao98' : {
		'Where_the_individual_ilet_is_connected_to' : None,
		'Who_has_built_your_toilet' : None,
		'Have_you_upgraded_yo_ng_individual_toilet' : None,
		'Cost_of_upgradation' : None,
		'Use_of_toilet' : None
	},
	'Note' : None,
	'Family_Photo' : None,
	'Toilet_Photo' : None,
	'__version__' : None,
	'meta' : {
		'instanceID' : None
	}
}

root_folder_path = ""

log_folder_path = "/home/softcorner/ShelterAssociates/src/scripts/Parag/Shelter Data Migration"

RA_excelFile = "/home/softcorner/ShelterAssociates/src/scripts/Parag/Shelter Data Migration/FilesToRead/RA.xls"
RA_mapped_excelFile = "/home/softcorner/ShelterAssociates/src/scripts/Parag/Shelter Data Migration/FilesToRead/MappedExcel_Pune/RA_Old_New_QuestionMapping_Parag.xlsx"

RHS_excelFile = "/home/softcorner/ShelterAssociates/src/scripts/Parag/Shelter Data Migration/FilesToRead/RHS.xls"
RHS_mapped_excelFile = "/home/softcorner/ShelterAssociates/src/scripts/Parag/Shelter Data Migration/FilesToRead/RHS_Old_New_QuestionMapping_Parag.xlsx"

FF_excelFile = "/home/softcorner/ShelterAssociates/src/scripts/Parag/Shelter Data Migration/FilesToRead/FF.xls"
FF_mapped_excelFile = "/home/softcorner/ShelterAssociates/src/scripts/Parag/Shelter Data Migration/FilesToRead/MappedExcel_Pune/FF_Old_New_QuestionMapping_Parag.xlsx"

ra_folder_path = "/home/softcorner/ShelterAssociates/src/scripts/Parag/Shelter Data Migration/xmloutput/RA/"
rhs_folder_path = "/home/softcorner/ShelterAssociates/src/scripts/Parag/Shelter Data Migration/xmloutput/RHS/"
ff_folder_path = "/home/softcorner/ShelterAssociates/src/scripts/Parag/Shelter Data Migration/xmloutput/FF/"

ra_xml_root = 'ap4nbe86zAhfHuePYwe8r7_bj4hnQo'
ra_xml_root_attr_id = 'ap4nbe86zAhfHuePYwe8r7'
ra_xml_root_attr_version = 'vqZZQxdvbcePcFQUM9KCxy'
ra_xml_formhub_uuid = '99605296640b4be591a6fe51ce42f853'

rhs_xml_root = 'aLaMKLSPygDyE5BmBzhUWz'
rhs_xml_root_attr_id = 'aLaMKLSPygDyE5BmBzhUWz'
rhs_xml_root_attr_version = 'vHPPjX2prBrs695YAWPmH7'
rhs_xml_formhub_uuid = '5284bf8321de44e78fd46f5629083975'

ff_xml_root = 'anR4hq5wSQoRTXoz7QHadB_v7ZuUIQ'
ff_xml_root_attr_id = 'anR4hq5wSQoRTXoz7QHadB'
ff_xml_root_attr_version = 'vpm89SWgdKWNfXRrSKq9eP'
ff_xml_formhub_uuid = 'c3c95bdbd7cb46f18a9ed899a448c84b'


survey_city_pune = 4
ra_survey_pune = 13
rhs_survey_pune = 46
rhs2_survey_pune = 17
ff_survey_pune = 18

survey_city_navi_mumabi = 47
ff_survey_navi_mumbai = 8

# queries to fetch data
qry_slum_list = "select distinct slum_id, slum_code from ray_survey_slumsurveymetadata where survey_id = %s order by slum_id"

# RA survey queries
qry_ra_survey_list = "select distinct f.object_id as slum_id from survey_fact f join slum_data_slum slum on slum.id = f.object_id \
join survey_survey s on s.id = f.survey_id join survey_project p on p.id = s.project_id \
where s.id = %s and p.id = %s and f.content_type_id = 26 order by f.object_id asc"

qry_ra_survey_slum_question_answer = "(select '-1' as question_id, to_char(min(f1.updated_on),'YYYY-MM-DD\"T\"HH24:MI:SS.MS+05:30') as answer from survey_fact f1 \
join survey_survey s1 on s1.id = f1.survey_id join survey_project p1 on p1.id = s1.project_id \
join survey_surveydesiredfact sdf1 on f1.desired_fact_id = sdf1.desired_fact_id and s1.id = sdf1.survey_id \
join slum_data_slum slum1 on slum1.id = f1.object_id \
where s1.id = %s and p1.id = %s and f1.content_type_id = 26 and slum1.id=%s)\
UNION All \
(select f.desired_fact_id as question_id, f.data as answer from survey_fact f \
join survey_survey s on s.id = f.survey_id join survey_project p on p.id = s.project_id \
join survey_surveydesiredfact sdf on f.desired_fact_id = sdf.desired_fact_id and s.id = sdf.survey_id \
join slum_data_slum slum on slum.id = f.object_id \
where s.id = %s and p.id = %s and f.content_type_id = 26 and slum.id=%s order by sdf.weight asc) "

qry_ra_survey_toilet_question_answer = "select toilet.id, f.desired_fact_id as question_id, f.data as answer from survey_fact f \
join survey_survey s on s.id = f.survey_id join survey_project p on p.id = s.project_id \
join survey_surveydesiredfact sdf on f.desired_fact_id = sdf.desired_fact_id and s.id = sdf.survey_id \
join slum_data_toilet toilet on toilet.id = f.object_id join slum_data_slum slum on slum.id = toilet.slum_id \
where s.id = %s and p.id = %s and f.content_type_id = 33 and slum.id=%s order by toilet_id, sdf.weight asc"

qry_ra_slum_household_count = "select count(id) from slum_data_household where slum_id=%s "

# RHS survey queries
qry_rhs_slum_household_survey_list = "select distinct household.slum_id, household.household_code from survey_fact f \
join slum_data_household household on household.id = f.object_id \
join survey_survey s on s.id = f.survey_id join survey_project p on p.id = s.project_id \
where s.id = %s and p.id = %s and f.content_type_id = 27 \
order by household.slum_id, household.household_code asc"

qry_rhs_survey_slum_household_question_answer = "(select household1.household_code, '-1' as question_id, to_char(min(f1.updated_on),'YYYY-MM-DD\"T\"HH24:MI:SS.MS+05:30') as answer from survey_fact f1 \
join survey_survey s1 on s1.id = f1.survey_id join survey_project p1 on p1.id = s1.project_id \
join survey_surveydesiredfact sdf1 on f1.desired_fact_id = sdf1.desired_fact_id and s1.id = sdf1.survey_id \
join slum_data_household household1 on household1.id = f1.object_id \
where s1.id = %s and p1.id = %s and f1.content_type_id = 27 \
 and household1.slum_id= %s group by household1.household_code) \
UNION All \
(select household.household_code, f.desired_fact_id as question_id, f.data as answer from survey_fact f \
join survey_survey s on s.id = f.survey_id join survey_project p on p.id = s.project_id \
join survey_surveydesiredfact sdf on f.desired_fact_id = sdf.desired_fact_id and s.id = sdf.survey_id \
join slum_data_household household on household.id = f.object_id \
where s.id = %s and p.id = %s and f.content_type_id = 27 and household.slum_id= %s order by household.household_code, sdf.weight asc)"


qry_rhs_common_slum_list = "(select distinct household.slum_id from survey_fact f \
join slum_data_household household on household.id = f.object_id \
join survey_survey s on s.id = f.survey_id join survey_project p on p.id = s.project_id \
where s.id = %s and p.id = %s and f.content_type_id = 27 order by household.slum_id asc) \
INTERSECT ALL \
(select distinct household.slum_id from survey_fact f \
join slum_data_household household on household.id = f.object_id \
join survey_survey s on s.id = f.survey_id join survey_project p on p.id = s.project_id \
where s.id = %s and p.id = %s and f.content_type_id = 27 \
order by household.slum_id asc)"

qry_rhs_master_slum_household_survey_list = "select distinct household.slum_id, household.household_code from survey_fact f \
join slum_data_household household on household.id = f.object_id \
join survey_survey s on s.id = f.survey_id join survey_project p on p.id = s.project_id \
where s.id = %s and p.id = %s and f.content_type_id = 27 \
order by household.slum_id, household.household_code asc"

qry_rhs_master_survey_slum_household_question_answer ="(select household1.household_code, '-1' as question_id, to_char(min(f1.updated_on),'YYYY-MM-DD\"T\"HH24:MI:SS.MS+05:30') as answer from survey_fact f1 \
join survey_survey s1 on s1.id = f1.survey_id join survey_project p1 on p1.id = s1.project_id \
join survey_surveydesiredfact sdf1 on f1.desired_fact_id = sdf1.desired_fact_id and s1.id = sdf1.survey_id \
join slum_data_household household1 on household1.id = f1.object_id \
where s1.id = %s and p1.id = %s and f1.content_type_id = 27 \
 and household1.slum_id= %s group by household1.household_code) \
UNION All \
(select household.household_code, f.desired_fact_id as question_id, f.data as answer from survey_fact f \
join survey_survey s on s.id = f.survey_id join survey_project p on p.id = s.project_id \
join survey_surveydesiredfact sdf on f.desired_fact_id = sdf.desired_fact_id and s.id = sdf.survey_id \
join slum_data_household household on household.id = f.object_id \
where s.id = %s and p.id = %s and f.content_type_id = 27 and household.slum_id= %s order by household.household_code, sdf.weight asc)"



# FF survey queries
 #start : for Pune
qry_ff_slum_household_survey_list = "select distinct household.slum_id, household.household_code from survey_fact f \
join slum_data_household household on household.id = f.object_id \
join survey_survey s on s.id = f.survey_id join survey_project p on p.id = s.project_id \
where s.id = %s and p.id = %s and f.content_type_id = 27 \
order by household.slum_id, household.household_code asc"

qry_ff_survey_slum_household_question_answer = "(select household1.household_code, '-1' as question_id, to_char(min(f1.updated_on),'YYYY-MM-DD\"T\"HH24:MI:SS.MS+05:30') as answer from survey_fact f1 \
join survey_survey s1 on s1.id = f1.survey_id join survey_project p1 on p1.id = s1.project_id \
join survey_surveydesiredfact sdf1 on f1.desired_fact_id = sdf1.desired_fact_id and s1.id = sdf1.survey_id \
join slum_data_household household1 on household1.id = f1.object_id \
where s1.id = %s and p1.id = %s and f1.content_type_id = 27 \
 and household1.slum_id= %s group by household1.household_code) \
UNION All \
(select household.household_code, f.desired_fact_id as question_id, f.data as answer from survey_fact f \
join survey_survey s on s.id = f.survey_id join survey_project p on p.id = s.project_id \
join survey_surveydesiredfact sdf on f.desired_fact_id = sdf.desired_fact_id and s.id = sdf.survey_id \
join slum_data_household household on household.id = f.object_id \
where s.id = %s and p.id = %s and f.content_type_id = 27 and household.slum_id= %s order by household.household_code, sdf.weight asc)"
#end : for Pune

qry_fact_option_text_list = "select code, description from survey_factoption where desired_fact_id=%s order by code asc"

"""
#start : for Navi Mumbai
qry_ff_slum_household_survey_list = "select distinct household.slum_id, household.household_code from survey_fact f \
join slum_data_household household on household.id = f.object_id \
join survey_survey s on s.id = f.survey_id join survey_project p on p.id = s.project_id \
where s.id = %s and p.id = %s and f.content_type_id = 27 \
order by household.slum_id, household.household_code asc"

qry_ff_survey_slum_household_question_answer = "(select household1.household_code, '-1' as question_id, to_char(min(f1.updated_on),'YYYY-MM-DD\"T\"HH24:MI:SS.MS+05:30') as answer from survey_fact f1 \
join survey_survey s1 on s1.id = f1.survey_id join survey_project p1 on p1.id = s1.project_id \
join survey_surveydesiredfact sdf1 on f1.desired_fact_id = sdf1.desired_fact_id and s1.id = sdf1.survey_id \
join slum_data_household household1 on household1.id = f1.object_id \
where s1.id = %s and p1.id = %s and f1.content_type_id = 27 \
 and household1.slum_id= %s group by household1.household_code) \
UNION All \
(select household.household_code, f.desired_fact_id as question_id, f.data as answer from survey_fact f \
join survey_survey s on s.id = f.survey_id join survey_project p on p.id = s.project_id \
join survey_surveydesiredfact sdf on f.desired_fact_id = sdf.desired_fact_id and s.id = sdf.survey_id \
join slum_data_household household on household.id = f.object_id \
where s.id = %s and p.id = %s and f.content_type_id = 27 and household.slum_id= %s order by household.household_code, sdf.weight asc)"
#end : for Navi Mumbai
"""
# get data from database
def fetch_db_records(query):
	global db_name
	global db_user
	global db_pwd
	global db_host
	global db_port

	conn = psycopg2.connect(database=db_name, user=db_user, password=db_pwd, host=db_host, port=db_port)
	#print("Opened database successfully")

	cur = conn.cursor()

	#print ("Query - ", query)
	cur.execute(query)

	#print ("Query executed Successfully")

	rows = cur.fetchall()
	#print ("Data fetched - Operation done successfully")
	#write_log("Data fetched : query  : " + query)

	conn.close()
	#print ("Closed database successfully")

	return rows;

#get single id list
def get_list_ids(query):
	data = []
	db_row = fetch_db_records(query)

	for row in db_row:
		data.append(row[0])

	return data;

# get dict for query
def get_question_answer(query):
	data = {}

	db_row = fetch_db_records(query)

	for row in db_row:
		key = row[0]
		val = row[1]

		if key not in data:
			data.setdefault(key, val)
		else:
			temp_val = data[key]
			if isinstance(temp_val, list):
				temp_val.append(val)
				data[key] = temp_val
			else:
				temp_lst = [temp_val, val]
				data[key] = temp_lst

	return data;

def get_toilet_question_answer(query):
	data = {}

	db_row = fetch_db_records(query)
	#print('data row -> ',db_row)
	for row in db_row:
		toilet = row[0]
		key = row[1]
		val = row[2]

		if toilet not in data:
			data.setdefault(toilet, {})

		toilet_dict = data[toilet]
		if key not in toilet_dict:
			toilet_dict.setdefault(key, val)
		else:
			temp_val = toilet_dict[key]

			if isinstance(temp_val, list):
				temp_val.append(val)
				toilet_dict[key] = temp_val
			else:
				temp_lst = [temp_val, val]
				toilet_dict[key] = temp_lst

	return data;

def get_household_wise_question_answer(query):
	data = {}

	db_row = fetch_db_records(query)
	#print('data row -> ',db_row)
	for row in db_row:
		household = row[0]
		key = row[1]
		val = row[2]

		if household not in data:
			data.setdefault(household, {})

		household_dict = data[household]
		if key not in household_dict:
			household_dict.setdefault(key, val)
		else:
			temp_val = household_dict[key]

			if isinstance(temp_val, list):
				temp_val.append(val)
				household_dict[key] = temp_val
			else:
				temp_lst = [temp_val, val]
				household_dict[key] = temp_lst

	return data;

def get_slum_code(query):
	return get_question_answer(query)

def get_household_count(query):
	data = 0
	db_row = fetch_db_records(query)

	for row in db_row:
		data = row[0]
		break;

	return data;

def get_household_survey(query):
	data = {}
	db_row = fetch_db_records(query)

	for row in db_row:
		household = row[0]
		survey = row[1]

		if household not in data:
			data.setdefault(household, [])

		data[household].append(survey)

	return data;

# function to read excel
def read_xml_excel(excelFile):
	global option_dict
	global city_ward_slum_dict

	global question_map_dict

	#open excel file
	#read question from sheet 1 into dict
	#map dict with option from sheet 2

	workbook = xlrd.open_workbook(excelFile)

	# List sheet names, and pull a sheet by name
	sheet_names = workbook.sheet_names()

	#print('Sheet Names', sheet_names)

	sheet_survey = workbook.sheet_by_index(0)
	sheet_choices = workbook.sheet_by_index(1)
	#sheet_settings = workbook.sheet_by_index(2)

	# read choice sheet to create option mapping dict
	for row in range(sheet_choices.nrows):
		if row != 0:
			key = sheet_choices.cell_value(row, 0)
			value = sheet_choices.cell_value(row, 1)

			#print('row value ', name)
			if key not in option_dict:
				option_dict.setdefault(key, [])

			option_dict[key].append(value)

			city = sheet_choices.cell_value(row, 3)
			if city.strip():
				if city not in city_ward_slum_dict:
					city_ward_slum_dict.setdefault(city, {})

				if value not in city_ward_slum_dict[city]:
					city_ward_slum_dict[city].setdefault(value, [])

			admin_ward = sheet_choices.cell_value(row, 4)

			if admin_ward.strip():
				for city in city_ward_slum_dict:
					if admin_ward in city_ward_slum_dict[city]:
						city_ward_slum_dict[city][admin_ward].append(value)
						break


	#print("dict - ", option_dict)
	#print("city_ward_slum_dict - ", city_ward_slum_dict)

	# read choice sheet to create option mapping dict
	for row in range(sheet_survey.nrows):
		if row != 0:
			name = sheet_survey.cell_value(row, 1)

			if name.strip() and not name.startswith("group_") and name not in question_map_dict:
				question_map_dict.setdefault(name, None)


	return;

def read_map_excel(excelFile):
	global question_map_dict
	global question_option_map_dict

	#open excel file
	#read mapping for quetion and new xml key
	#create mapping dict

	workbook = openpyxl.load_workbook(excelFile)

	# List sheet names, and pull a sheet by name
	sheet_names = workbook.get_sheet_names()

	#print('Sheet Names', sheet_names)

	sheet_old_new_question_mapping = workbook.worksheets[0]
	sheet_old_new_option_mapping = workbook.worksheets[2]

	for row in sheet_old_new_question_mapping.iter_rows(row_offset=1):
		row_data  = []
		for cell in row:
			row_data.append(cell.value)

		question_id = row_data[5]
		if not (question_id is None):
			dict_key = row_data[1]

			question_map_dict[dict_key] = question_id

	#print("dict question map - ", question_map_dict)

	# set option mapping dict for all qustion mapped
	for row in sheet_old_new_option_mapping.iter_rows(row_offset=1):
		row_data  = []
		for cell in row:
			row_data.append(cell.value)

		question_id = row_data[8]

		if not (question_id is None):
			old_option = row_data[9]

			if not (old_option is None):
				new_option = row_data[2]

				if question_id not in question_option_map_dict:
					question_option_map_dict.setdefault(question_id, {})

				# in case multiple old option map to single option
				if isinstance(old_option, int):
					question_option_map_dict[question_id].setdefault(old_option, new_option)
				else:
					old_option_list = old_option.split(',')
					for option_id in old_option_list:
						question_option_map_dict[question_id].setdefault(int(option_id), new_option)

	#print("dict optioin map - ", question_option_map_dict)

	return;

def get_answer(xml_key, fact_dict):
	global question_map_dict
	global question_option_map_dict
	answer = None

	if xml_key in question_map_dict:
		fact_id = question_map_dict[xml_key]
		if (fact_id is None):
			return None

		#get answer as return
		if fact_id in fact_dict:
			answer = fact_dict[fact_id]

			#check fact has options
			if fact_id in question_option_map_dict:
				#check if answer is list - this is in case of multi select option
				if isinstance(answer, list):
					temp_answer = None
					for ans in answer:
						if int(ans) in question_option_map_dict[fact_id]:
							temp_ans = question_option_map_dict[fact_id][int(ans)]
							if not (temp_ans is None):
								if temp_answer is None:
									temp_answer = str('')
								temp_answer = temp_answer + str(temp_ans) + ' '

					if not(temp_answer is None):
						answer = temp_answer.strip()
				else:
					if answer:
						answer_option = int(answer)
						if answer_option in question_option_map_dict[fact_id]:
							answer = question_option_map_dict[fact_id][answer_option]

	return answer;

def get_name_id(xml_key, answer_text):
	global option_dict

	answer = None

	#print('answer_text => ',answer_text)

	option_list = option_dict[xml_key]
	#print('option_list => ',option_list)

	if answer_text:
		name_list = answer_text.split(',')

		if not name_list:
			name_list = answer_text.split('/')

		#print('name_list => ',name_list)

		if name_list:
			for name in name_list:
				for option in option_list:
					#print(xml_key+"    "+name +"   "+option)
					if option.lower() == name.lower():
						answer = option


	return answer;

def get_admin_ward(slum_code):
	admin_ward = None
	global city_ward_slum_dict

	for city, admin_ward_dict in iter(city_ward_slum_dict.items()):
		if admin_ward_dict:
			for ward, slum in iter(admin_ward_dict.items()):
				if slum_code in slum:
					admin_ward = ward
					break;
		if not (admin_ward is None):
			break;
	return admin_ward;

def get_city_id(admin_ward):
	city_id = None
	global city_ward_slum_dict

	for city, admin_ward_dict in iter(city_ward_slum_dict.items()):
		if admin_ward_dict:
			if admin_ward in admin_ward_dict:
				city_id = city
				break;

	return city_id;

	return;

def get_formatted_data(date_string):
	date_after_format = None
	date_converted = None

	date_format = ['%d/%m/%Y', '%d/%m/%y', '%d.%m.%Y', '%d.%m.%y', '%d-%m-%Y', '%d %b %Y', '%dth %b %Y', '%dth %B %Y', '%dth %B %Y.', '%dth %B, %Y', '%dth %b. %Y',
					'%dst %b %Y', '%dnd %b %Y', '%drd %b %Y', '%dth %b %y', '%d/%b/%Y', '%dth %B %Y', '%dnd %B %Y', '%dst %B %Y', '%dnd %B %y', '%drd %B %Y', '%B %Y',
					'%d/%m/%Y']

	for format in date_format:
		if date_converted is None:
			try:
				date_converted = datetime.datetime.strptime(date_string, format)
				break
			except:
				date_converted = None
				pass

	if date_converted:
		date_after_format = date_converted.strftime('%Y-%m-%d') #'%Y-%m-%dT%H:%M:%S-%z'

	return date_after_format;

def convert_area_from_square_meters(area_sq_m):
	area = 0

	if area_sq_m:
		area_sq_m = area_sq_m.lower()

		unformatted_area = area_sq_m.replace('sq.','').replace('m.','')

		unformatted_area =  unformatted_area.replace('m','')

		unformatted_area =  unformatted_area.replace('ts','')

		unformatted_area =  unformatted_area.replace('sq','')

		unformatted_area = unformatted_area.replace(' ','')
		unformatted_area = unformatted_area.replace(',','')

		try:
			#print('converted area=>'+unformatted_area+"  float => "+str(float(unformatted_area)))
			area = int(float(unformatted_area.strip()))
		except:
			write_log('unformatted area=>'+unformatted_area)
			#print('converted area=>'+unformatted_area)
			pass

		#print("final area => "+str(area))

	return area;

def get_approximat_huts(slum_id, answer):
	huts_count = 0
	option = {1:75, 2:225, 3:400, 4:750, 5:1000}

	huts_count = get_household_count(qry_ra_slum_household_count % slum_id)

	if huts_count == 0:
		if answer:
			huts_count = option[int(answer)]

	return huts_count;

def get_rhs_area_in_squar_feet(answer_sq_m):
	area = 0

	if answer_sq_m:
		try:
			area = int(answer_sq_m)
		except Exception:
			if '*' in answer_sq_m:
				area_size = answer_sq_m.split('*')
				area = int(area_size[0]) * int(area_size[1])
			elif ',' in answer_sq_m:
				area = int(answer_sq_m.replace(',',''))
			elif '/' in answer_sq_m:
				area = int(answer_sq_m.replace('/',''))
			elif isinstance(answer_sq_m, list):
				area = int(answer_sq_m[0])
			elif '`' in answer_sq_m:
				area = int(answer_sq_m.replace('`',''))
			elif 'sq' in answer_sq_m.lower():
				area_sq_m = answer_sq_m.lower()

				unformatted_area = area_sq_m.replace('sq','').replace('.ft','')

				area = int(unformatted_area.strip())
			else:
				raise Exception

	return area;

def get_rhs_family_member_count(answer_count):
	count = None

	if answer_count:
		try:
			count = int(answer_count)
		except Exception:
			if ',' in answer_count:
				count = int(answer_count.replace(',',''))
			elif '*' in answer_count:
				count = int(answer_count.replace('*',''))
			elif '?' in answer_count:
				count = int(answer_count.replace('?',''))
			elif '/' in answer_count:
				count = int(answer_count.replace('/',''))
			elif isinstance(answer_count, list):
				count = int(answer_count[0])
			else:
				raise Exception

	return count;

def get_option_text(option_list, answer):
	option_text = ''

	if answer:
		if isinstance(answer, list):
			temp_option = ''
			for id in answer:
				if id:
					temp_option += option_list[id] + ','

			option_text = temp_option[:-1]
		else:
			option_text = option_list[answer]

	return option_text;

def create_xml_string(xml_dict, repeat_dict, xml_root, xml_root_attr_id, xml_root_attr_version):
	xml_string = dicttoxml.dicttoxml(xml_dict, attr_type=False, custom_root=xml_root)
	#print(xml_string)
	#print("\n")
	root = ET.fromstring(xml_string)
	root.set('id', xml_root_attr_id)
	root.set('version', xml_root_attr_version)

	#repeat_dict = {'group_te3dx03' : { 'append_index' : 1, 'list' : toilet_info}}
	if repeat_dict:
		for key, val in repeat_dict.items():
			if val['list']:
				sub_ele = root.find(key)
				index = val['append_index']
				# create xml to be appened and append
				for sub_xml_dict in val['list']:
					sub_xml_string = dicttoxml.dicttoxml(sub_xml_dict, attr_type=False, root=False)
					#print('\n sub xml - %s -- '%index ,sub_xml_string)

					sub_root = ET.fromstring(sub_xml_string)

					sub_ele.insert(index, sub_root)
					index = index+1

	xml_string = ET.tostring(root, encoding="utf8", method='xml')
	#print('\n final xml -- ', xml_string)
	#write_log('created xml string to write')

	return root;

def create_xml_file(xml_root, filename, folderpath):
	file = filename + ".xml"
	xml_file = folderpath + file;

	directory = os.path.dirname(xml_file)

	if not os.path.exists(directory):
		os.makedirs(directory)

	xml_tree = ET.ElementTree(xml_root)

	xml_tree.write(xml_file, xml_declaration=True, encoding='utf-8', method="xml")

	log_msg = "created xml file : " + xml_file
	#write_log(log_msg)
	#print(log_msg)

	return;


def write_log(msg):
	global log_folder_path

	log_file = log_folder_path + 'log_' + str(strftime("%d_%m_%Y", gmtime())) + '.txt'

	directory = os.path.dirname(log_file)

	if not os.path.exists(directory):
		os.makedirs(directory)

	cur_datetime = strftime("%Y-%m-%d %H:%M:%S", gmtime())

	msg_str = cur_datetime + "\t\t" + msg + "\n"

	filehandle = open(log_file, "a")

	filehandle.write(msg_str)

	filehandle.close()

	return;

# Logic for convert to xml
# get list of ids (survey in case of RA, household and survey for RHS and FF)
# loop through ids to create xml
# fetch fact/answers to question to create xml
# match xml key with question no and its answer
# 	if answer is text take anser as it is
# 	if answer is options then from answer in db get new id of xml and set into xml
# set few key values mannually like guid and others
#
# required mapping
# 1. xml key and DB question mapping
# 2. DB option mapping with new xml options
# 3. xml keys into dict to create xml
# 4. mapping for slum - admin ward - city from DB to set into xml

def create_ra_xml(project_id, survey_id, excelfile, mapexcelfile, folderpath):
	# get list of all servey to create xml for each survey
	global question_map_dict
	global question_option_map_dict
	global option_dict

	global city_ward_slum_dict

	global qry_slum_list
	global qry_ra_survey_list
	global qry_ra_survey_question_answer

	global gl_ra_xml_dict

	global ra_xml_root
	global ra_xml_root_attr_id
	global ra_xml_root_attr_version
	global ra_xml_formhub_uuid

	global ra_folder_path

	name_mismatch_records = {
		'name_of_surveyor_s_who_checke' : {'count':0, 'slum':[]},
		'name_of_surveyor_s_who_collec' : {'count':0, 'slum':[]},
		'name_of_surveyor_s_who_took_t' : {'count':0, 'slum':[]}
	}
	unprocess_records = {}

	write_log("Start : Log for RA Survey for each slum ")

	#read old xls file city - ward - slum mapping
	read_xml_excel(excelfile)
	#print("Read excel file")
	write_log("Read excel file " + excelfile)

	#print(city_ward_slum_dict)

	#read map xlsx file for question, option mapping
	read_map_excel(mapexcelfile)
	#print("Read mapped excel file")
	write_log("Read mapped excel file" + mapexcelfile)

	# get slum code list
	slum_code_list = get_slum_code(qry_slum_list % survey_id)
	#print("fatch slum code")
	#print(slum_code_list)
	write_log("fatch slum code")

	survey_list = get_list_ids(qry_ra_survey_list % (survey_id, project_id))
	#print("fetch survey list")
	#print(survey_list)
	write_log("fetch survey list")

	fail = 0
	success = 0

	for slum in survey_list:
		try:
			print("proocessing data for slum - ", slum)
			write_log("proocessing data for slum id : "+ str(slum))
			toilet_info = []

			slum_fact = get_question_answer(qry_ra_survey_slum_question_answer % (survey_id, project_id, slum, survey_id, project_id, slum))
			toilet_fact = get_toilet_question_answer(qry_ra_survey_toilet_question_answer % (survey_id, project_id, slum))

			fact = {}
			fact.update(slum_fact)
			#fact.update(toilet_fact)

			slum_code = slum_code_list[slum]
			if slum_code:
				#print('slum_fact - ', slum_fact)
				#print('\n')
				#print('toilet_fact - ', toilet_fact)
				#print('\n')
				#print('question answer', fact)

				admin_ward = get_admin_ward(slum_code)
				city = get_city_id(admin_ward)

				ra_xml_dict = copy.deepcopy(gl_ra_xml_dict)

				group_tb3th42_dict = copy.deepcopy(ra_xml_dict['group_te3dx03']['group_tb3th42'])
				del ra_xml_dict['group_te3dx03']['group_tb3th42']

				del ra_xml_dict['group_zl6oo94']['group_wb1hp47']['Date_of_declaration']


				toilet_block_details = group_tb3th42_dict.copy()

				#print('slum_code : %s  admin_ward : %s  city : %s' % (slum_code, admin_ward, city))

				#print('toilet_block_details : ',toilet_block_details)


				ra_xml_dict['formhub']['uuid'] = ra_xml_formhub_uuid

				ra_xml_dict['start'] = get_answer('start', fact)
				ra_xml_dict['end'] = get_answer('end', fact)

				#Administration section
				date_of_survey = get_answer('date_of_survey', fact) #check date format
				if date_of_survey:
					#print("date_of_survey ="+date_of_survey)
					ra_xml_dict['group_ws5ux48']['date_of_survey'] = get_formatted_data(date_of_survey)

				name_of_surveyor_s_who_checke = get_answer('name_of_surveyor_s_who_checke', fact)
				name_of_surveyor_1 = get_name_id('sf02f44', name_of_surveyor_s_who_checke)
				if name_of_surveyor_1 is None:
					name_mismatch_records['name_of_surveyor_s_who_checke']['count'] += 1
					name_mismatch_records['name_of_surveyor_s_who_checke']['slum'].append(slum)
				else:
					ra_xml_dict['group_ws5ux48']['name_of_surveyor_s_who_checke'] = name_of_surveyor_1

				name_of_surveyor_s_who_collec = get_answer('name_of_surveyor_s_who_collec', fact)
				name_of_surveyor_2 = get_name_id('qs59t66', name_of_surveyor_s_who_collec)
				if name_of_surveyor_2 is None:
					name_mismatch_records['name_of_surveyor_s_who_collec']['count'] += 1
					name_mismatch_records['name_of_surveyor_s_who_collec']['slum'].append(slum)
				else:
					ra_xml_dict['group_ws5ux48']['name_of_surveyor_s_who_collec'] = name_of_surveyor_2

				name_of_surveyor_s_who_took_t = get_answer('name_of_surveyor_s_who_took_t', fact)
				name_of_surveyor_3 = get_name_id('gf4ac05', name_of_surveyor_s_who_took_t)
				if name_of_surveyor_3 is None:
					name_mismatch_records['name_of_surveyor_s_who_took_t']['count'] += 1
					name_mismatch_records['name_of_surveyor_s_who_took_t']['slum'].append(slum)
				else:
					ra_xml_dict['group_ws5ux48']['name_of_surveyor_s_who_took_t'] = name_of_surveyor_3

				#print('process - Administration section')
				#write_log('process - Administration section')

				#General information - Part A
				ra_xml_dict['group_zl6oo94']['group_uj8eg07']['city'] = city
				ra_xml_dict['group_zl6oo94']['group_uj8eg07']['admin_ward'] = admin_ward
				ra_xml_dict['group_zl6oo94']['group_uj8eg07']['slum_name'] = slum_code
				ra_xml_dict['group_zl6oo94']['group_uj8eg07']['survey_sector_number'] = get_answer('survey_sector_number', fact)
				ra_xml_dict['group_zl6oo94']['group_uj8eg07']['landmark'] = get_answer('landmark', fact)

				#print('process - General information - Part A')
				#write_log('process - General information - Part A')

				#General information - Part B
				ra_xml_dict['group_zl6oo94']['group_wb1hp47']['year_established_according_to'] = get_answer('year_established_according_to', fact)
				legal_status = get_answer('legal_status', fact)

				if legal_status:
					ra_xml_dict['group_zl6oo94']['group_wb1hp47']['legal_status'] = legal_status
					if legal_status != '01':
						ra_xml_dict['group_zl6oo94']['group_wb1hp47']['Date_of_declaration'] = get_answer('Date_of_declaration', fact)
					#else:
						#del ra_xml_dict['group_zl6oo94']['group_wb1hp47']['Date_of_declaration']

				land_owner = get_answer('land_owner', fact)
				if land_owner:
					ra_xml_dict['group_zl6oo94']['group_wb1hp47']['land_owner'] = land_owner

				development_plan_reservation_t = get_answer('development_plan_reservation_t', fact)
				if development_plan_reservation_t:
					ra_xml_dict['group_zl6oo94']['group_wb1hp47']['development_plan_reservation_t'] = development_plan_reservation_t

				ra_xml_dict['group_zl6oo94']['group_wb1hp47']['development_plan_reservation'] = get_answer('development_plan_reservation', fact)

				approximate_area_of_the_settle = get_answer('approximate_area_of_the_settle', fact)
				ra_xml_dict['group_zl6oo94']['group_wb1hp47']['approximate_area_of_the_settle'] = convert_area_from_square_meters(approximate_area_of_the_settle)

				number_of_huts_in_settlement = get_approximat_huts(slum, get_answer('number_of_huts_in_settlement', fact))
				ra_xml_dict['group_zl6oo94']['group_wb1hp47']['number_of_huts_in_settlement'] = number_of_huts_in_settlement

				location = get_answer('location', fact)
				if location:
					ra_xml_dict['group_zl6oo94']['group_wb1hp47']['location'] = location

				topography = get_answer('topography', fact)
				if topography:
					ra_xml_dict['group_zl6oo94']['group_wb1hp47']['topography'] = topography

				ra_xml_dict['group_zl6oo94']['group_wb1hp47']['describe_the_slum'] = get_answer('describe_the_slum', fact)

				#print('process - General information - Part B')
				#write_log('process - General information - Part B')

				#Toilet Information - Status of Defecation
				status_of_defecation = get_answer('status_of_defecation', fact)
				if status_of_defecation:
					ra_xml_dict['group_te3dx03']['group_ul75r92']['status_of_defecation'] = status_of_defecation

				number_of_community_toilet_blo = get_answer('number_of_community_toilet_blo', fact)
				ra_xml_dict['group_te3dx03']['group_ul75r92']['number_of_community_toilet_blo'] = number_of_community_toilet_blo

				if number_of_community_toilet_blo and int(number_of_community_toilet_blo) > 0:
					no_toilet = int(number_of_community_toilet_blo)
					no_actual_toilet = len(toilet_fact)
					if no_toilet != no_actual_toilet:
						ra_xml_dict['group_te3dx03']['group_ul75r92']['number_of_community_toilet_blo'] = no_actual_toilet

					ra_xml_dict['group_te3dx03']['group_ul75r92']['number_of_pay_and_use_CTBs'] = int(get_answer('number_of_pay_and_use_CTBs', fact))

					# loop for no of toilet
					for toilet_id in toilet_fact.keys():
						#Toilet Information - Toilet Block Details
						toilet_details_dict = {}

						is_the_CTB_in_use = get_answer('is_the_CTB_in_use', toilet_fact[toilet_id])
						if is_the_CTB_in_use:
							toilet_details_dict['is_the_CTB_in_use'] = is_the_CTB_in_use

							if is_the_CTB_in_use == '04':
								toilet_details_dict['fee_for_use_of_ctb_per_family'] = int(get_answer('fee_for_use_of_ctb_per_family', toilet_fact[toilet_id]))
								toilet_details_dict['cost_of_pay_and_use_toilet_pe'] = int(get_answer('cost_of_pay_and_use_toilet_pe', toilet_fact[toilet_id]))
								toilet_details_dict['ctb_gender_usage'] = get_answer('ctb_gender_usage', toilet_fact[toilet_id])

								ctb_gender_usage = get_answer('ctb_gender_usage', toilet_fact[toilet_id])
								if ctb_gender_usage:
									if ctb_gender_usage == '03':
										toilet_details_dict['total_number_of_mixed_seats_al'] = 0 #get_answer('total_number_of_mixed_seats_al', toilet_fact[toilet_id])

										number_of_mixed_seats_allotted = get_answer('number_of_mixed_seats_allotted', toilet_fact[toilet_id])
										toilet_details_dict['number_of_mixed_seats_allotted'] = 0 #get_answer('number_of_mixed_seats_allotted', toilet_fact[toilet_id])

										#if number_of_mixed_seats_allotted is not None and int(number_of_mixed_seats_allotted) > 0:
										#	toilet_details_dict['the_reason_for_the_mixed_seats'] = None #get_answer('the_reason_for_the_mixed_seats', toilet_fact[toilet_id])
									elif ctb_gender_usage == '01' or ctb_gender_usage == '04':
										toilet_details_dict['number_of_seats_allotted_to_me'] = int(get_answer('number_of_seats_allotted_to_me', toilet_fact[toilet_id]))

										number_of_seats_allotted_to_me_001 = get_answer('number_of_seats_allotted_to_me_001', toilet_fact[toilet_id])
										toilet_details_dict['number_of_seats_allotted_to_me_001'] = int(number_of_seats_allotted_to_me_001)

										if int(number_of_seats_allotted_to_me_001) > 0:
											the_reason_for_men_not_using_t = get_answer('the_reason_for_men_not_using_t', toilet_fact[toilet_id])
											if the_reason_for_men_not_using_t:
												toilet_details_dict['the_reason_for_men_not_using_t'] = the_reason_for_men_not_using_t
									elif ctb_gender_usage == '02' or ctb_gender_usage == '04':
										toilet_details_dict['number_of_seats_allotted_to_wo'] = int(get_answer('number_of_seats_allotted_to_wo', toilet_fact[toilet_id]))

										number_of_seats_allotted_to_wo_001 = get_answer('number_of_seats_allotted_to_wo_001', toilet_fact[toilet_id])
										toilet_details_dict['number_of_seats_allotted_to_wo_001'] = int(number_of_seats_allotted_to_wo_001)

										if int(number_of_seats_allotted_to_wo_001) > 0:
											the_reason_for_women_not_using = get_answer('the_reason_for_women_not_using', toilet_fact[toilet_id])
											if the_reason_for_women_not_using:
												toilet_details_dict['the_reason_for_women_not_using'] = the_reason_for_women_not_using

								is_the_ctb_available_at_night = get_answer('is_the_ctb_available_at_night', toilet_fact[toilet_id])
								if is_the_ctb_available_at_night:
									toilet_details_dict['is_the_ctb_available_at_night'] = is_the_ctb_available_at_night

								ctb_maintenance_provided_by = get_answer('ctb_maintenance_provided_by', toilet_fact[toilet_id])
								if ctb_maintenance_provided_by:
									toilet_details_dict['ctb_maintenance_provided_by'] = ctb_maintenance_provided_by

								condition_of_ctb_structure = get_answer('condition_of_ctb_structure', toilet_fact[toilet_id])
								if condition_of_ctb_structure:
									toilet_details_dict['condition_of_ctb_structure'] = condition_of_ctb_structure

								toilet_details_dict['out_of_total_seats_no_of_pans_in_good_condition'] = None #get_answer('out_of_total_seats_no_of_pans_in_good_condition', toilet_fact[toilet_id])
								toilet_details_dict['out_of_total_seats_no_of_doors_in_good_condition'] = None #get_answer('out_of_total_seats_no_of_doors_in_good_condition', toilet_fact[toilet_id])
								toilet_details_dict['out_of_total_seats_no_of_seats_where_electricity_is_available'] = None #get_answer('out_of_total_seats_no_of_seats_where_electricity_is_available', toilet_fact[toilet_id])
								toilet_details_dict['out_of_total_seats_no_of_seats_where_tiles_on_wall_are_in_good_condition'] = None #get_answer('out_of_total_seats_no_of_seats_where_tiles_on_wall_are_in_good_condition', toilet_fact[toilet_id])
								toilet_details_dict['out_of_total_seats_no_of_seats_where_tiles_on_floor_are_in_good_condition'] = None #get_answer('out_of_total_seats_no_of_seats_where_tiles_on_floor_are_in_good_condition', toilet_fact[toilet_id])

								frequency_of_ctb_cleaning_by_U = get_answer('frequency_of_ctb_cleaning_by_U', toilet_fact[toilet_id])
								if frequency_of_ctb_cleaning_by_U:
									toilet_details_dict['frequency_of_ctb_cleaning_by_U'] = frequency_of_ctb_cleaning_by_U

								does_the_ulb_ngo_communty_use = get_answer('does_the_ulb_ngo_communty_use', toilet_fact[toilet_id])
								if does_the_ulb_ngo_communty_use:
									toilet_details_dict['does_the_ulb_ngo_communty_use'] = does_the_ulb_ngo_communty_use

								cleanliness_of_the_ctb = get_answer('cleanliness_of_the_ctb', toilet_fact[toilet_id])
								if cleanliness_of_the_ctb:
									toilet_details_dict['cleanliness_of_the_ctb'] = cleanliness_of_the_ctb

								is_there_a_caretaker_for_the_C = get_answer('is_there_a_caretaker_for_the_C', toilet_fact[toilet_id])
								if is_there_a_caretaker_for_the_C:
									toilet_details_dict['is_there_a_cretaker_for_the_C'] = is_there_a_caretaker_for_the_C

								type_of_water_supply_in_ctb = get_answer('type_of_water_supply_in_ctb', toilet_fact[toilet_id])
								if type_of_water_supply_in_ctb:
									toilet_details_dict['type_of_water_supply_in_ctb'] = type_of_water_supply_in_ctb

									if type_of_water_supply_in_ctb == '02':
										capacity_of_ctb_water_tank_in = get_answer('capacity_of_ctb_water_tank_in', toilet_fact[toilet_id])
										if capacity_of_ctb_water_tank_in:
											toilet_details_dict['capacity_of_ctb_water_tank_in'] = capacity_of_ctb_water_tank_in

								litres_of_water_used_by_commun = get_answer('litres_of_water_used_by_commun', toilet_fact[toilet_id])
								if litres_of_water_used_by_commun:
									toilet_details_dict['litres_of_water_used_by_commun'] = litres_of_water_used_by_commun

								availability_of_water_in_the_t = get_answer('availability_of_water_in_the_t', toilet_fact[toilet_id])
								if availability_of_water_in_the_t:
									toilet_details_dict['availability_of_water_in_the_t'] = availability_of_water_in_the_t

								availability_of_electricity_in = get_answer('availability_of_electricity_in', toilet_fact[toilet_id])
								if availability_of_electricity_in:
									toilet_details_dict['availability_of_electricity_in'] = availability_of_electricity_in

								availability_of_electricity_in_001 = get_answer('availability_of_electricity_in_001', toilet_fact[toilet_id])
								if availability_of_electricity_in_001:
									toilet_details_dict['availability_of_electricity_in_001'] = availability_of_electricity_in_001

								facility_in_the_toilet_block_f = get_answer('facility_in_the_toilet_block_f', toilet_fact[toilet_id])
								if facility_in_the_toilet_block_f:
									toilet_details_dict['facility_in_the_toilet_block_f'] = facility_in_the_toilet_block_f

									if facility_in_the_toilet_block_f == '02':
										condition_of_facility_for_chil = get_answer('condition_of_facility_for_chil', toilet_fact[toilet_id])
										if condition_of_facility_for_chil:
											toilet_details_dict['condition_of_facility_for_chil'] = condition_of_facility_for_chil

								sewage_disposal_system = get_answer('sewage_disposal_system', toilet_fact[toilet_id])
								if sewage_disposal_system:
									toilet_details_dict['sewage_disposal_system'] = sewage_disposal_system

								distance_to_nearest_ulb_sewer = get_answer('distance_to_nearest_ulb_sewer', toilet_fact[toilet_id])
								if distance_to_nearest_ulb_sewer:
									toilet_details_dict['distance_to_nearest_ulb_sewer'] = distance_to_nearest_ulb_sewer

							# add into list
							toilet_info.append({'group_tb3th42': toilet_details_dict})
				else:
					del ra_xml_dict['group_te3dx03']['group_ul75r92']['number_of_pay_and_use_CTBs']

				ra_xml_dict['group_te3dx03']['toilet_comment'] = get_answer('toilet_comment', fact)

				#print('process - Toilet Information - Status of Defecation')
				#write_log('process - Toilet Information - Status of Defecation')

				#Water Information
				Total_number_of_standposts_in_ = get_answer('Total_number_of_standposts_in_', fact)

				if Total_number_of_standposts_in_:
					ra_xml_dict['group_zj8tc43']['Total_number_of_standposts_in_'] = int(Total_number_of_standposts_in_)

					if  int(Total_number_of_standposts_in_) > 0:
						ra_xml_dict['group_zj8tc43']['total_number_of_taps_in_use_n'] = get_answer('total_number_of_taps_in_use_n', fact)

				Total_number_of_standposts_NOT = get_answer('Total_number_of_standposts_NOT', fact)

				if Total_number_of_standposts_NOT:
					ra_xml_dict['group_zj8tc43']['Total_number_of_standposts_NOT'] = Total_number_of_standposts_NOT

					if (not(Total_number_of_standposts_in_ is None) and int(Total_number_of_standposts_in_) > 0) or int(Total_number_of_standposts_NOT) > 0:
						ra_xml_dict['group_zj8tc43']['total_number_of_taps_in_use_n_001'] = get_answer('total_number_of_taps_in_use_n_001', fact)

				total_number_of_handpumps_in_u = get_answer('total_number_of_handpumps_in_u', fact)
				ra_xml_dict['group_zj8tc43']['total_number_of_handpumps_in_u'] = int(total_number_of_handpumps_in_u) if total_number_of_handpumps_in_u else 0

				total_number_of_handpumps_in_u_001 = get_answer('total_number_of_handpumps_in_u_001', fact)
				ra_xml_dict['group_zj8tc43']['total_number_of_handpumps_in_u_001'] = int(total_number_of_handpumps_in_u_001) if total_number_of_handpumps_in_u_001 else 0

				alternative_source_of_water = get_answer('alternative_source_of_water', fact)
				if alternative_source_of_water:
					ra_xml_dict['group_zj8tc43']['alternative_source_of_water'] = alternative_source_of_water

				availability_of_water = get_answer('availability_of_water', fact)
				if availability_of_water:
					ra_xml_dict['group_zj8tc43']['availability_of_water'] = availability_of_water

				pressure_of_water_in_the_syste = get_answer('pressure_of_water_in_the_syste', fact)
				if pressure_of_water_in_the_syste:
					ra_xml_dict['group_zj8tc43']['pressure_of_water_in_the_syste'] = pressure_of_water_in_the_syste

				coverage_of_wateracross_settle = get_answer('coverage_of_wateracross_settle', fact)
				if coverage_of_wateracross_settle:
					ra_xml_dict['group_zj8tc43']['coverage_of_wateracross_settle'] = coverage_of_wateracross_settle

				quality_of_water_in_the_system = get_answer('quality_of_water_in_the_system', fact)
				if quality_of_water_in_the_system:
					ra_xml_dict['group_zj8tc43']['quality_of_water_in_the_system'] = quality_of_water_in_the_system

				ra_xml_dict['group_zj8tc43']['water_supply_comment'] = get_answer('water_supply_comment', fact)

				#print('process - Water Information')
				#write_log('process - Water Information')

				#Waste Management Information
				total_number_of_waste_containe = get_answer('total_number_of_waste_containe', fact)
				ra_xml_dict['group_ks0wh10']['total_number_of_waste_containe'] = int(total_number_of_waste_containe) if total_number_of_waste_containe else 0

				facility_of_waste_collection = get_answer('facility_of_waste_collection', fact)
				if facility_of_waste_collection:
					ra_xml_dict['group_ks0wh10']['facility_of_waste_collection'] = facility_of_waste_collection

					#if facility_of_waste_collection == '02':
						#ra_xml_dict['group_ks0wh10']['frequency_of_waste_collection'] = None #get_answer('frequency_of_waste_collection', fact)
						#ra_xml_dict['group_ks0wh10']['coverage_of_waste_collection_a'] = None #get_answer('coverage_of_waste_collection_a', fact)
					#elif facility_of_waste_collection == '03':
						#ra_xml_dict['group_ks0wh10']['frequency_of_waste_collection__002'] = None #get_answer('frequency_of_waste_collection__002', fact)
						#ra_xml_dict['group_ks0wh10']['coverage_of_waste_collection_a_001'] = None #get_answer('coverage_of_waste_collection_a_001', fact)
					#elif facility_of_waste_collection == '04':
						#ra_xml_dict['group_ks0wh10']['frequency_of_waste_collection_001'] = None #get_answer('frequency_of_waste_collection_001', fact)
						#ra_xml_dict['group_ks0wh10']['coverage_of_waste_collection_a_002'] = None #get_answer('coverage_of_waste_collection_a_002', fact)
					#elif facility_of_waste_collection == '05':
						#ra_xml_dict['group_ks0wh10']['frequency_of_waste_collection_'] = None #get_answer('frequency_of_waste_collection_', fact)
						#ra_xml_dict['group_ks0wh10']['coverage_of_waste_collection_a_003'] = None #get_answer('coverage_of_waste_collection_a_003', fact)
					#elif facility_of_waste_collection == '06':
						#ra_xml_dict['group_ks0wh10']['frequency_of_waste_collection__001'] = None #get_answer('frequency_of_waste_collection__001', fact)

				where_are_the_communty_open_du = get_answer('where_are_the_communty_open_du', fact)
				if where_are_the_communty_open_du:
					ra_xml_dict['group_ks0wh10']['where_are_the_communty_open_du'] = where_are_the_communty_open_du

				do_the_member_of_community_dep = get_answer('do_the_member_of_community_dep', fact)
				if do_the_member_of_community_dep:
					ra_xml_dict['group_ks0wh10']['do_the_member_of_community_dep'] = do_the_member_of_community_dep

				ra_xml_dict['group_ks0wh10']['Waste_management_comments'] = get_answer('Waste_management_comments', fact)

				#print('process - Waste Management Information')
				#write_log('process - Waste Management Information')

				#Drainage Information
				presence_of_drains_within_the = get_answer('presence_of_drains_within_the', fact)
				if presence_of_drains_within_the:
					ra_xml_dict['group_kk5gz02']['presence_of_drains_within_the'] = presence_of_drains_within_the

					if presence_of_drains_within_the == '02':
						coverage_of_drains_across_the = get_answer('coverage_of_drains_across_the', fact)
						if coverage_of_drains_across_the:
							ra_xml_dict['group_kk5gz02']['coverage_of_drains_across_the'] = coverage_of_drains_across_the

						do_the_drains_get_blocked = get_answer('do_the_drains_get_blocked', fact)
						if do_the_drains_get_blocked:
							ra_xml_dict['group_kk5gz02']['do_the_drains_get_blocked'] = do_the_drains_get_blocked

						is_the_drainage_gradient_adequ = get_answer('is_the_drainage_gradient_adequ', fact)
						if is_the_drainage_gradient_adequ:
							ra_xml_dict['group_kk5gz02']['is_the_drainage_gradient_adequ'] = is_the_drainage_gradient_adequ

						#ra_xml_dict['group_kk5gz02']['diameter_of_ulb_sewer_line_acr'] = None #get_answer('diameter_of_ulb_sewer_line_acr', fact)

				ra_xml_dict['group_kk5gz02']['drainage_comment'] = get_answer('drainage_comment', fact)

				#print('process - Drainage Information')
				#write_log('process - Drainage Information')

				#Gutter Information
				#Presence_of_gutter = get_answer('Presence_of_gutter', fact)
				#if Presence_of_gutter:
					#ra_xml_dict['group_bv7hf31']['Presence_of_gutter'] = None #Presence_of_gutter

					#if Presence_of_gutter == '02':
						#ra_xml_dict['group_bv7hf31']['type_of_gutter_within_the_sett'] = None #get_answer('type_of_gutter_within_the_sett', fact)
						#ra_xml_dict['group_bv7hf31']['coverage_of_gutter'] = None #get_answer('coverage_of_gutter', fact)
						#ra_xml_dict['group_bv7hf31']['are_gutter_covered'] = None #get_answer('are_gutter_covered', fact)
						#ra_xml_dict['group_bv7hf31']['do_gutters_flood'] = None #get_answer('do_gutters_flood', fact)
						#ra_xml_dict['group_bv7hf31']['do_gutter_get_choked'] = None #get_answer('do_gutter_get_choked', fact)
						#ra_xml_dict['group_bv7hf31']['is_gutter_gradient_adequate'] = None #get_answer('is_gutter_gradient_adequate', fact)

				#ra_xml_dict['group_bv7hf31']['comments_on_gutter'] = None #get_answer('comments_on_gutter', fact)

				#print('process - Gutter Information')
				#write_log('process - Gutter Information')

				#Roads and Access Information
				presence_of_roads_within_the_s = get_answer('presence_of_roads_within_the_s', fact)
				if presence_of_roads_within_the_s:
					ra_xml_dict['group_xy9hz30']['presence_of_roads_within_the_s'] = presence_of_roads_within_the_s

					if presence_of_roads_within_the_s == '02':
						type_of_roads_within_the_settl = get_answer('type_of_roads_within_the_settl', fact)
						if type_of_roads_within_the_settl:
							ra_xml_dict['group_xy9hz30']['type_of_roads_within_the_settl'] = type_of_roads_within_the_settl

							if type_of_roads_within_the_settl == '03':
								coverage_of_pucca_road_across = get_answer('coverage_of_pucca_road_across', fact)
								if coverage_of_pucca_road_across:
									ra_xml_dict['group_xy9hz30']['coverage_of_pucca_road_across'] = coverage_of_pucca_road_across

						finish_of_the_road = get_answer('finish_of_the_road', fact)
						if finish_of_the_road:
							ra_xml_dict['group_xy9hz30']['finish_of_the_road'] = finish_of_the_road

						average_width_of_internal_road = get_answer('average_width_of_internal_road', fact)
						if average_width_of_internal_road:
							ra_xml_dict['group_xy9hz30']['average_width_of_internal_road'] = average_width_of_internal_road

						average_width_of_arterial_road = get_answer('average_width_of_arterial_road', fact)
						if average_width_of_arterial_road:
							ra_xml_dict['group_xy9hz30']['average_width_of_arterial_road'] = average_width_of_arterial_road

						point_of_vehicular_access_to_t = get_answer('point_of_vehicular_access_to_t', fact)
						if point_of_vehicular_access_to_t:
							ra_xml_dict['group_xy9hz30']['point_of_vehicular_access_to_t'] = point_of_vehicular_access_to_t

						is_the_settlement_below_or_abo = get_answer('is_the_settlement_below_or_abo', fact)
						if is_the_settlement_below_or_abo:
							ra_xml_dict['group_xy9hz30']['is_the_settlement_below_or_abo'] = is_the_settlement_below_or_abo

						are_the_huts_below_or_above_th = get_answer('are_the_huts_below_or_above_th', fact)
						if are_the_huts_below_or_above_th:
							ra_xml_dict['group_xy9hz30']['are_the_huts_below_or_above_th'] = are_the_huts_below_or_above_th

				ra_xml_dict['group_xy9hz30']['road_and_access_comment'] = get_answer('road_and_access_comment', fact)

				#print('process - Roads and Access Information')
				#write_log('process - Roads and Access Information')

				ra_xml_dict['__version__'] = ra_xml_root_attr_version

				ra_xml_dict['meta']['instanceID'] = 'uuid:' + str(uuid.uuid4())


				# code to write data into XML file
				repeat_dict = {'group_te3dx03' : { 'append_index' : 1, 'list' : toilet_info}}
				xml_root = create_xml_string(ra_xml_dict, repeat_dict, ra_xml_root, ra_xml_root_attr_id, ra_xml_root_attr_version)

				file_name = 'RA_Survey_Slum_Id_' + str(slum)
				create_xml_file(xml_root, file_name, ra_folder_path)

				success += 1
			else:
				# write log that slum code is not found for slum id
				write_log('slum code is not found for slum id '+str(slum))
				unprocess_records[str(slum)] = 'slum code is not found when mapped'
				fail += 1

			#print ('ra data - ', ra_xml_dict)
			#print ('ra toilate data - ', toilet_info)

			del ra_xml_dict

			#break;
		except Exception as ex:
			exception_log = 'Exception occurred for slum id '+str(slum)+' \t  exception : '+ str(ex) +' \t  traceback : '+ traceback.format_exc()
			unprocess_records[str(slum)] = str(ex)

			fail += 1
			write_log(exception_log)

			#break;
			pass

	if unprocess_records:
		write_log('List of slum for which unable to create xml')
		write_log('slum_id \t exception')
		for slum_id, error_msg in unprocess_records.items():
			write_log(slum_id+' \t\t'+error_msg)

	write_log('End : Log for RA Survey for each slum \n')
	print("End processing")

	total = len(survey_list)
	result_log = 'total records : '+str(total) + ' \t fail to process : '+str(fail) + ' \t success : '+str(success)

	print(result_log)
	write_log(result_log)

	name_mismatch_records_log = 'Name of surveyor not match -  \t '
	name_mismatch_records_log += ' name_of_surveyor_s_who_checke : '+str(name_mismatch_records['name_of_surveyor_s_who_checke']['count']) + ' \t '
	name_mismatch_records_log += ' name_of_surveyor_s_who_collec : '+str(name_mismatch_records['name_of_surveyor_s_who_collec']['count']) + ' \t '
	name_mismatch_records_log += ' name_of_surveyor_s_who_took_t : '+str(name_mismatch_records['name_of_surveyor_s_who_took_t']['count']) + ' \t '
	print(name_mismatch_records_log)
	write_log(name_mismatch_records_log)

	return;

def create_rhs_xml(project_id, survey_id, survey_id2, excelfile, mapexcelfile, folderpath):
	# get list of all servey to create xml for each survey
	global question_map_dict
	global question_option_map_dict
	global option_dict

	global city_ward_slum_dict

	global qry_slum_list
	global qry_rhs_slum_household_survey_list
	global qry_rhs_survey_slum_household_question_answer

	global qry_rhs_common_slum_list
	global qry_rhs_master_slum_household_survey_list
	global qry_rhs_master_survey_slum_household_question_answer

	global gl_rhs_xml_dict

	global rhs_xml_root
	global rhs_xml_root_attr_id
	global rhs_xml_root_attr_version
	global rhs_xml_formhub_uuid

	global rhs_folder_path

	unprocess_records = {}

	write_log("Start : Log for RHS Survey for per household in each slum ")

	#read old xls file city - ward - slum mapping
	read_xml_excel(excelfile)
	#print("Read excel file")
	write_log("Read excel file " + excelfile)

	#print(city_ward_slum_dict)

	#read map xlsx file for question, option mapping
	read_map_excel(mapexcelfile)
	#print("Read mapped excel file")
	write_log("Read mapped excel file" + mapexcelfile)

	# get slum code list
	slum_code_list = {}
	slum_code1_list = get_slum_code(qry_slum_list % survey_id)

	if survey_id2:
		slum_code2_list = get_slum_code(qry_slum_list % survey_id2)

		slum_code2_list.update(slum_code1_list)

		slum_code_list.update(slum_code2_list)
	else:
		slum_code_list.update(slum_code1_list)

	#print("fatch slum code")
	write_log("fatch slum code")

	rhs_group = {'master':None, 'New':None}

	common_slum_id_list = get_list_ids(qry_rhs_common_slum_list  % (survey_id, project_id, survey_id2, project_id))
	write_log("fatch common slum in both data for RHS -- " + (', '.join(str(x) for x in common_slum_id_list)))

	#print('common_slum_id_list=> ', common_slum_id_list)

	master_slum_household_list = get_household_survey(qry_rhs_master_slum_household_survey_list % (survey_id2, project_id))
	#print('master_slum_household_list before - ',master_slum_household_list.keys())

	for slum_id in common_slum_id_list:
		del master_slum_household_list[slum_id]

	#print("fetch master slum household list")
	#print('master_slum_household_list  ',master_slum_household_list.keys())

	rhs_group['master']= master_slum_household_list

	new_slum_household_list = get_household_survey(qry_rhs_slum_household_survey_list % (survey_id, project_id))
	#print("fetch slum household list")
	#print(new_slum_household_list)
	write_log("fetch household slum survey list")

	rhs_group['New'] = new_slum_household_list

	fail = 0
	success = 0
	total_process_house = 0

	for rhs_key, slum_household_list in rhs_group.items():
		#check key value
		#print('rhs_key = ', rhs_key)
		#print('slum_household_list == ',slum_household_list)

		for slum, household_list in slum_household_list.items():
			print("proocessing data for slum - ", slum)
			write_log("proocessing data for slum - "+str(slum))

			unprocess_records.setdefault(str(slum), [])

			try:
				slum_code = None
				slum_code = slum_code_list[slum]
			except:
				pass

			if slum_code:
				admin_ward = get_admin_ward(slum_code)
				city = get_city_id(admin_ward)

				#print('slum_code : %s  admin_ward : %s  city : %s' % (slum_code, admin_ward, city))

				qry_rhs_question_answer = ''
				if rhs_key == 'master':
					qry_rhs_question_answer = (qry_rhs_master_survey_slum_household_question_answer  % (survey_id2, project_id, slum, survey_id2, project_id, slum))
				else:
					qry_rhs_question_answer = (qry_rhs_survey_slum_household_question_answer  % (survey_id, project_id, slum, survey_id, project_id, slum))

				household_fact = get_household_wise_question_answer(qry_rhs_question_answer)
				#print(household_fact)
				total_process_house += len(household_fact)

				for household in household_list:
					try:
						print("proocessing data for household - %s in slum - %s" % (household, slum))
						#write_log("proocessing data for household - %s in slum - %s" % (household, str(slum)))

						fact = household_fact[household]

						#print('question answer', fact)

						rhs_xml_dict = copy.deepcopy(gl_rhs_xml_dict)

						rhs_xml_dict['formhub']['uuid'] = rhs_xml_formhub_uuid

						rhs_xml_dict['start'] = get_answer('start', fact)
						rhs_xml_dict['end'] = get_answer('end', fact)

						#Administrative Information
						rhs_xml_dict['group_ce0hf58']['city'] = city
						rhs_xml_dict['group_ce0hf58']['admin_ward'] = admin_ward
						rhs_xml_dict['group_ce0hf58']['slum_name'] = slum_code

						date_of_rhs = get_answer('date_of_rhs', fact)
						if date_of_rhs:
							rhs_xml_dict['group_ce0hf58']['date_of_rhs'] = get_formatted_data(date_of_rhs)

						rhs_xml_dict['group_ce0hf58']['name_of_surveyor_who_collected_rhs_data'] = get_answer('name_of_surveyor_who_collected_rhs_data', fact)
						rhs_xml_dict['group_ce0hf58']['house_no'] = get_answer('house_no', fact)

						Type_of_structure_occupancy = get_answer('Type_of_structure_occupancy', fact)
						if Type_of_structure_occupancy:
							rhs_xml_dict['group_ce0hf58']['Type_of_structure_occupancy'] = Type_of_structure_occupancy

						#print('process - Administrative Information')
						#write_log('process - Administrative Information')

						#Household Information - Personal Information
						if Type_of_structure_occupancy == '01' or Type_of_structure_occupancy == '03':
							rhs_xml_dict['group_ye18c77']['group_ud4em45']['what_is_the_full_name_of_the_family_head_'] = get_answer('what_is_the_full_name_of_the_family_head_', fact)
							rhs_xml_dict['group_ye18c77']['group_ud4em45']['mobile_number'] = get_answer('mobile_number', fact)
							rhs_xml_dict['group_ye18c77']['group_ud4em45']['adhar_card_number'] = get_answer('adhar_card_number', fact)

						#print('process - Household Information - Personal Information')
						#write_log('process - Household Information - Personal Information')

						#Household Information - General Information
						if Type_of_structure_occupancy == '01':
							what_is_the_structure_of_the_house = get_answer('what_is_the_structure_of_the_house', fact)
							if what_is_the_structure_of_the_house:
								rhs_xml_dict['group_ye18c77']['group_yw8pj39']['what_is_the_structure_of_the_house'] = what_is_the_structure_of_the_house

							what_is_the_ownership_status_of_the_house = get_answer('what_is_the_ownership_status_of_the_house', fact)
							if what_is_the_ownership_status_of_the_house:
								rhs_xml_dict['group_ye18c77']['group_yw8pj39']['what_is_the_ownership_status_of_the_house'] = what_is_the_ownership_status_of_the_house

							number_of_family_members = get_answer('number_of_family_members', fact)
							try:
								rhs_xml_dict['group_ye18c77']['group_yw8pj39']['number_of_family_members'] = get_rhs_family_member_count(number_of_family_members)
							except:
								#unprocess_records[str(slum)].append([str(household), "unable to process number of family member for answer =>"+(number_of_family_members if not isinstance(number_of_family_members, list) else ','.join(number_of_family_members))])
								pass

							Do_you_have_a_girl_child_under = get_answer('Do_you_have_a_girl_child_under', fact)
							if Do_you_have_a_girl_child_under:
								rhs_xml_dict['group_ye18c77']['group_yw8pj39']['Do_you_have_a_girl_child_under'] = Do_you_have_a_girl_child_under

								if Do_you_have_a_girl_child_under == '01':
									rhs_xml_dict['group_ye18c77']['group_yw8pj39']['if_yes_how_many_'] = int(get_answer('if_yes_how_many_', fact))

							house_area_in_sq_ft = get_answer('house_area_in_sq_ft', fact)
							try:
								rhs_xml_dict['group_ye18c77']['group_yw8pj39']['house_area_in_sq_ft'] = get_rhs_area_in_squar_feet(house_area_in_sq_ft)
							except:
								#unprocess_records[str(slum)].append([str(household), "unable to process house area in sq ft for answer =>"+(house_area_in_sq_ft if not isinstance(house_area_in_sq_ft, list) else ','.join(house_area_in_sq_ft))])
								pass

							Current_place_of_defecation_toilet = get_answer('Current_place_of_defecation_toilet', fact)
							if Current_place_of_defecation_toilet:
								rhs_xml_dict['group_ye18c77']['group_yw8pj39']['Current_place_of_defecation_toilet'] = Current_place_of_defecation_toilet

							does_any_member_of_your_family_go_for_open_defecation_ = get_answer('does_any_member_of_your_family_go_for_open_defecation_', fact)
							if does_any_member_of_your_family_go_for_open_defecation_:
								rhs_xml_dict['group_ye18c77']['group_yw8pj39']['does_any_member_of_your_family_go_for_open_defecation_'] = does_any_member_of_your_family_go_for_open_defecation_

							if Current_place_of_defecation_toilet and (Current_place_of_defecation_toilet == '01' or Current_place_of_defecation_toilet == '02'):
								where_the_individual_toilet_is_connected_to_ = get_answer('where_the_individual_toilet_is_connected_to_', fact)
								if where_the_individual_toilet_is_connected_to_:
									rhs_xml_dict['group_ye18c77']['group_yw8pj39']['where_the_individual_toilet_is_connected_to_'] = where_the_individual_toilet_is_connected_to_

							type_of_water_connection = get_answer('type_of_water_connection', fact)
							if type_of_water_connection:
								rhs_xml_dict['group_ye18c77']['group_yw8pj39']['type_of_water_connection'] = type_of_water_connection

							facility_of_waste_collection = get_answer('facility_of_waste_collection', fact)
							if facility_of_waste_collection:
								rhs_xml_dict['group_ye18c77']['group_yw8pj39']['facility_of_waste_collection'] = facility_of_waste_collection

							if Current_place_of_defecation_toilet and (Current_place_of_defecation_toilet != '01' or Current_place_of_defecation_toilet != '02'):
								Are_you_interested_in_individu = get_answer('Are_you_interested_in_individu', fact)
								if Are_you_interested_in_individu:
									rhs_xml_dict['group_ye18c77']['group_yw8pj39']['Are_you_interested_in_individu'] = Are_you_interested_in_individu

									if Are_you_interested_in_individu == '01':
										if_yes_why_ = get_answer('if_yes_why_', fact)
										if if_yes_why_:
											rhs_xml_dict['group_ye18c77']['group_yw8pj39']['if_yes_why_'] = if_yes_why_

									if Are_you_interested_in_individu == '02':
										if_no_why_ = get_answer('if_no_why_', fact)
										if if_no_why_:
											rhs_xml_dict['group_ye18c77']['group_yw8pj39']['if_no_why_'] = if_no_why_

									if Are_you_interested_in_individu == '01' and city != 3789:
										type_of_toilet_preference = get_answer('type_of_toilet_preference', fact)
										if type_of_toilet_preference:
											rhs_xml_dict['group_ye18c77']['group_yw8pj39']['type_of_toilet_preference'] = type_of_toilet_preference

							if Current_place_of_defecation_toilet and Current_place_of_defecation_toilet != '01':
								Have_you_applied_for_indiviual = get_answer('Have_you_applied_for_indiviual', fact)
								if Have_you_applied_for_indiviual:
									rhs_xml_dict['group_ye18c77']['grup_yw8pj39']['Have_you_applied_for_indiviual'] = Have_you_applied_for_indiviual

									if Have_you_applied_for_indiviual == '01':
										How_many_installements_have_yo = get_answer('How_many_installements_have_yo', fact)
										if How_many_installements_have_yo:
											rhs_xml_dict['group_ye18c77']['group_yw8pj39']['How_many_installements_have_yo'] = How_many_installements_have_yo

											if How_many_installements_have_yo == '02' or How_many_installements_have_yo == '03':
												rhs_xml_dict['group_ye18c77']['group_yw8pj39']['when_did_you_receive_the_first_installment_date'] = get_answer('when_did_you_receive_the_first_installment_date', fact)

											if How_many_installements_have_yo == '03':
												rhs_xml_dict['group_ye18c77']['group_yw8pj39']['when_did_you_receive_the_second_installment_date'] = get_answer('when_did_you_receive_the_second_installment_date', fact)

										what_is_the_status_of_toilet_under_sbm_ = get_answer('what_is_the_status_of_toilet_under_sbm_', fact)
										if what_is_the_status_of_toilet_under_sbm_:
											rhs_xml_dict['group_ye18c77']['group_yw8pj39']['what_is_the_status_of_toilet_under_sbm_'] = what_is_the_status_of_toilet_under_sbm_

							Does_any_family_members_has_co = get_answer('Does_any_family_members_has_co', fact)
							if Does_any_family_members_has_co:
								rhs_xml_dict['group_ye18c77']['group_yw8pj39']['Does_any_family_members_has_co'] = Does_any_family_members_has_co

						#print('process - Household Information - General Information')
						#write_log('process - Household Information - General Information')

						rhs_xml_dict['__version__'] = rhs_xml_root_attr_version

						rhs_xml_dict['meta']['instanceID'] = 'uuid:' + str(uuid.uuid4())


						# code to write data into XML file
						repeat_dict = {}
						xml_root = create_xml_string(rhs_xml_dict, repeat_dict, rhs_xml_root, rhs_xml_root_attr_id, rhs_xml_root_attr_version)

						file_name = 'RHS_Survey_Slum_Id_' + str(slum) + '_House_code_' + household
						create_xml_file(xml_root, file_name, rhs_folder_path + "slum_" +str(slum) + "/")

						success += 1

						#print ('rhs data - ', rhs_xml_dict)

						del rhs_xml_dict

						#break;
					except Exception as ex:
						exception_log = 'Exception occurred for household id ' +str(household) + ' of slum id ' + str(slum) + ' \t  exception : '+ str(ex) +' \t  traceback : '+ traceback.format_exc()
						unprocess_records[str(slum)].append([str(household), str(ex)])

						fail += 1
						write_log(exception_log)

						#break;
						pass
				#break;
			else:
				# write log that slum code is not found for slum id
				write_log('slum code is not found for slum id '+str(slum))
				unprocess_records[str(slum)].append([None, 'slum code is not found when mapped'])
				fail += 1
		#break;

	if unprocess_records:
		write_log('List of slum and household for which unable to create xml')
		write_log('slum_id \t household_code \t exception')
		for slum_id, error_lst in unprocess_records.items():
			for error in error_lst:
				#print('error ', error[0], '    msg ',error[1])
				write_log(slum_id+' \t\t' + (error[0]+' \t' if error[0] else ' \t\t') +' \t\t\t' + error[1])

	write_log('End : Log for RHS Survey for slum \n')
	print("End processing")

	total_slum = len(rhs_group['master']) + len(rhs_group['New'])
	total_house = sum(len(v1) for v1 in iter(rhs_group['master'].values())) + sum(len(v2) for v2 in iter(rhs_group['New'].values()))

	result_log = 'total slum records : '+str(total_slum) +'\t total house records in all slum : '+str(total_house)
	print(result_log)
	write_log(result_log)

	result_log2 = 'process house records in all slum : '+str(total_process_house) + ' \t fail to process : '+str(fail) + ' \t success : '+str(success)
	print(result_log2)
	write_log(result_log2)

	return;

def create_ff_xml(project_id, survey_id, excelfile, mapexcelfile, folderpath):
	# get list of all servey to create xml for each survey
	global question_map_dict
	global question_option_map_dict
	global option_dict

	global city_ward_slum_dict

	global qry_slum_list
	global qry_ff_slum_household_survey_list
	global qry_ff_survey_slum_household_question_answer

	global qry_fact_option_text_list

	global gl_ff_xml_dict

	global ff_xml_root
	global ff_xml_root_attr_id
	global ff_xml_root_attr_version
	global ff_xml_formhub_uuid

	global ff_folder_path

	unprocess_records = {}

	write_log("Start : Log for FF Survey for per household in each slum ")

	#read old xls file city - ward - slum mapping
	read_xml_excel(excelfile)
	#print("Read excel file")
	write_log("Read excel file " + excelfile)

	#print(city_ward_slum_dict)

	#read map xlsx file for question, option mapping
	read_map_excel(mapexcelfile)
	#print("Read mapped excel file")
	write_log("Read mapped excel file" + mapexcelfile)

	# get slum code list
	slum_code_list = get_slum_code(qry_slum_list % survey_id)
	#print("fatch slum code")
	#print(slum_code_list)
	write_log("fatch slum code")

	slum_household_list = get_household_survey(qry_ff_slum_household_survey_list % (survey_id, project_id))
	#print("fetch slum household list")
	#print(new_slum_household_list)
	write_log("fetch household slum survey list")

	occupation_option_list = get_question_answer(qry_fact_option_text_list % 434)
	disability_option_list = get_question_answer(qry_fact_option_text_list % 432)

	#print(occupation_option_list)
	#print(disability_option_list)

	fail = 0
	success = 0
	total_process_house = 0

	for slum, household_list in slum_household_list.items():
		print("proocessing data for slum - ", slum)
		write_log("proocessing data for slum - "+str(slum))

		unprocess_records.setdefault(str(slum), [])

		try:
			slum_code = None
			slum_code = slum_code_list[slum]
		except:
			pass

		if slum_code: #and slum in []:
			admin_ward = get_admin_ward(slum_code)
			city = get_city_id(admin_ward)

			#print('slum_code : %s  admin_ward : %s  city : %s' % (slum_code, admin_ward, city))

			household_fact = get_household_wise_question_answer(qry_ff_survey_slum_household_question_answer % (survey_id, project_id, slum, survey_id, project_id, slum))
			#print(household_fact)
			total_process_house += len(household_fact)

			for household in household_list:
				try:
					print("proocessing data for household - %s in slum - %s" % (household, slum))
					#write_log("proocessing data for household - %s in slum - %s" % (household, str(slum)))

					fact = household_fact[household]

					#print('question answer', fact)

					ff_xml_dict = copy.deepcopy(gl_ff_xml_dict)

					ff_xml_dict['formhub']['uuid'] = ff_xml_formhub_uuid

					#Slum Information
					ff_xml_dict['group_vq77l17']['city'] = city
					ff_xml_dict['group_vq77l17']['admin_ward'] = admin_ward
					ff_xml_dict['group_vq77l17']['slum_name'] = slum_code
					ff_xml_dict['group_vq77l17']['Settlement_address'] = get_answer('Settlement_address', fact)

					Household_number = get_answer('Household_number', fact)
					try:
						ff_xml_dict['group_vq77l17']['Household_number'] = int(Household_number)
					except:
						#unprocess_records[str(slum)].append([str(household), "unable to process house number for answer =>"+(Household_number if not isinstance(Household_number, list) else ','.join(Household_number))])
						pass

					#print('process - Slum Information')
					#write_log('process - Slum Information')

					#Family Information
					ff_xml_dict['group_oh4zf84']['Name_of_the_family_head'] = get_answer('Name_of_the_family_head', fact)
					ff_xml_dict['group_oh4zf84']['Name_of_Native_villa_district_and_state'] = get_answer('Name_of_Native_villa_district_and_state', fact)
					ff_xml_dict['group_oh4zf84']['Duration_of_stay_in_the_city'] = get_answer('Duration_of_stay_in_the_city', fact)
					ff_xml_dict['group_oh4zf84']['Duration_of_stay_in_s_current_settlement'] = get_answer('Duration_of_stay_in_s_current_settlement', fact)

					Type_of_house = get_answer('Type_of_house', fact)
					if Type_of_house:
						ff_xml_dict['group_oh4zf84']['Type_of_house'] = Type_of_house

					Ownership_status = get_answer('Ownership_status', fact)
					if Ownership_status:
						ff_xml_dict['group_oh4zf84']['Ownership_status'] = Ownership_status

					#print('process - Family Information')
					#write_log('process - Family Information')

					#Family Members Information
					Total_family_members = get_answer('Total_family_members', fact)
					try:
						ff_xml_dict['group_im2th52']['Total_family_members'] = int(Total_family_members)
					except:
						#unprocess_records[str(slum)].append([str(household), "unable to process number of male member for answer =>"+(Total_family_members if Total_family_members else 'NoneTYpe')])
						pass

					Number_of_Male_members = get_answer('Number_of_Male_members', fact)
					try:
						ff_xml_dict['group_im2th52']['Number_of_Male_members'] = int(Number_of_Male_members)
					except:
						#unprocess_records[str(slum)].append([str(household), "unable to process number of male member for answer =>"+(Number_of_Male_members if Number_of_Male_members else 'NoneTYpe')])
						pass

					Number_of_Female_members = get_answer('Number_of_Female_members', fact)
					try:
						ff_xml_dict['group_im2th52']['Number_of_Female_members'] = int(Number_of_Female_members)
					except:
						#unprocess_records[str(slum)].append([str(household), "unable to process number of male member for answer =>"+(Number_of_Female_members if Number_of_Female_members else 'NoneTYpe')])
						pass

					Number_of_Children_under_5_years_of_age = get_answer('Number_of_Children_under_5_years_of_age', fact)
					try:
						ff_xml_dict['group_im2th52']['Number_of_Children_under_5_years_of_age'] = int(Number_of_Children_under_5_years_of_age)
					except:
						#unprocess_records[str(slum)].append([str(household), "unable to process number of children under 5 years of age for answer =>"+(Number_of_Children_under_5_years_of_age)])
						pass

					Number_of_members_over_60_years_of_age = get_answer('Number_of_members_over_60_years_of_age', fact)
					try:
						ff_xml_dict['group_im2th52']['Number_of_members_over_60_years_of_age'] = int(Number_of_members_over_60_years_of_age)
					except:
						#unprocess_records[str(slum)].append([str(household), "unable to process number of member over 60 years of age for answer =>"+(Number_of_members_over_60_years_of_age)])
						pass

					Number_of_Girl_children_between_0_18_yrs = get_answer('Number_of_Girl_children_between_0_18_yrs', fact)
					try:
						ff_xml_dict['group_im2th52']['Number_of_Girl_children_between_0_18_yrs'] = int(Number_of_Girl_children_between_0_18_yrs)
					except:
						#unprocess_records[str(slum)].append([str(household), "unable to process number of girl children between 0-18 of age for answer =>"+(Number_of_Girl_children_between_0_18_yrs if Number_of_Girl_children_between_0_18_yrs else 'NoneTYpe')])
						pass

					Number_of_disabled_members = get_answer('Number_of_disabled_members', fact)
					try:
						ff_xml_dict['group_im2th52']['Number_of_disabled_members'] = int(Number_of_disabled_members)

						if Number_of_disabled_members > 0:
							ff_xml_dict['group_im2th52']['If_yes_specify_type_of_disability'] = get_option_text(disability_option_list, get_answer('If_yes_specify_type_of_disability', fact))
					except:
						#unprocess_records[str(slum)].append([str(household), "unable to process number of disable member for answer =>"+(Number_of_disabled_members if Number_of_disabled_members else 'NoneTYpe')])
						pass

					Number_of_earning_members = get_answer('Number_of_earning_members', fact)
					try:
						ff_xml_dict['group_im2th52']['Number_of_earning_members'] = int(Number_of_earning_members)
					except:
						#unprocess_records[str(slum)].append([str(household), "unable to process number of disable member for answer =>"+(Number_of_earning_members if Number_of_earning_members else 'NoneTYpe')])
						pass

					ff_xml_dict['group_im2th52']['Occupation_s_of_earning_members'] = get_option_text(occupation_option_list, get_answer('Occupation_s_of_earning_members', fact))

					ff_xml_dict['group_im2th52']['Approximate_monthly_family_income'] = get_answer('Approximate_monthly_family_income', fact)

					#print('process - Family Members Information')
					#write_log('process - Family Members Information')

					#Toilet Information
					Where_the_individual_ilet_is_connected_to = get_answer('Where_the_individual_ilet_is_connected_to', fact)
					if Where_the_individual_ilet_is_connected_to:
						ff_xml_dict['group_ne3ao98']['Where_the_individual_ilet_is_connected_to'] = Where_the_individual_ilet_is_connected_to

					Who_has_built_your_toilet = get_answer('Who_has_built_your_toilet', fact)
					if Who_has_built_your_toilet:
						ff_xml_dict['group_ne3ao98']['Who_has_built_your_toilet'] = Who_has_built_your_toilet

					Have_you_upgraded_yo_ng_individual_toilet = get_answer('Have_you_upgraded_yo_ng_individual_toilet', fact)
					if Have_you_upgraded_yo_ng_individual_toilet:
						ff_xml_dict['group_ne3ao98']['Have_you_upgraded_yo_ng_individual_toilet'] = Have_you_upgraded_yo_ng_individual_toilet

					ff_xml_dict['group_ne3ao98']['Cost_of_upgradation'] = get_answer('Cost_of_upgradation', fact)

					Use_of_toilet = get_answer('Use_of_toilet', fact)
					if Use_of_toilet:
						ff_xml_dict['group_ne3ao98']['Use_of_toilet'] = Use_of_toilet

					#print('process - Toilet Information')
					#write_log('process - Toilet Information')

					ff_xml_dict['Note'] = get_answer('Note', fact)
					ff_xml_dict['Family_Photo'] = None #get_answer('Family_Photo', fact)
					ff_xml_dict['Toilet_Photo'] = None #get_answer('Toilet_Photo', fact)

					ff_xml_dict['__version__'] = ff_xml_root_attr_version

					ff_xml_dict['meta']['instanceID'] = 'uuid:' + str(uuid.uuid4())


					# code to write data into XML file
					repeat_dict = {}
					xml_root = create_xml_string(ff_xml_dict, repeat_dict, ff_xml_root, ff_xml_root_attr_id, ff_xml_root_attr_version)

					file_name = 'FF_Survey_Slum_Id_' + str(slum) + '_House_code_' + household
					create_xml_file(xml_root, file_name, ff_folder_path + "slum_" +str(slum) + "/")

					success += 1

					#print ('ff data - ', ff_xml_dict)

					del ff_xml_dict

					#break;
				except Exception as ex:
					exception_log = 'Exception occurred for household id ' +str(household) + ' of slum id ' + str(slum) + ' \t  exception : '+ str(ex) +' \t  traceback : '+ traceback.format_exc()
					unprocess_records[str(slum)].append([str(household), str(ex)])

					fail += 1
					write_log(exception_log)

					#break;
					pass
			#break;
		else:
			# write log that slum code is not found for slum id
			write_log('slum code is not found for slum id '+str(slum))
			unprocess_records[str(slum)].append([None, 'slum code is not found when mapped'])
			fail += 1


	if unprocess_records:
		write_log('List of slum and household for which unable to create xml')
		write_log('slum_id \t household_code \t exception')
		for slum_id, error_lst in unprocess_records.items():
			for error in error_lst:
				#print('error ', error[0], '    msg ',error[1])
				write_log(slum_id+' \t\t' + (error[0]+' \t' if error[0] else ' \t\t') +' \t\t\t' + error[1])

	write_log('End : Log for FF Survey for slum \n')
	print("End processing")

	total_slum = len(slum_household_list)
	total_house = sum(len(v1) for v1 in iter(slum_household_list.values()))

	result_log = 'total slum records : '+str(total_slum) +'\t total house records in all slum : '+str(total_house)
	print(result_log)
	write_log(result_log)

	result_log2 = 'process house records in all slum : '+str(total_process_house) + ' \t fail to process : '+str(fail) + ' \t success : '+str(success)
	print(result_log2)
	write_log(result_log2)

	return;


# call to create xml files for survey
create_ra_xml(survey_city_pune, ra_survey_pune, RA_excelFile, RA_mapped_excelFile, ra_folder_path)

#create_rhs_xml(survey_city_pune, rhs_survey_pune, rhs2_survey_pune, RHS_excelFile, RHS_mapped_excelFile, rhs_folder_path)

#create_ff_xml(survey_city_pune, ff_survey_pune, FF_excelFile, FF_mapped_excelFile, ff_folder_path)
